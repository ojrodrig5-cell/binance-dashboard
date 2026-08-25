"""
Binance Top Gainers Tracker (FUTUROS USDⓈ-M)
-----------------------------------------------
Consulta la API pública de Binance Futures (no requiere login ni API key),
obtiene los contratos perpetuos con mayor % de ganancia en las últimas 24h,
guarda un registro en Excel cada vez que se ejecuta, y genera
una gráfica con la evolución del top gainer.

Requisitos:
    pip install requests openpyxl matplotlib plyer

Uso:
    python binance_tracker.py

Se recomienda programar este script para correr cada hora usando
el Programador de tareas de Windows (Task Scheduler). Instrucciones
al final de este archivo.
"""

import requests
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import PatternFill, Font
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
import os
import shutil
import subprocess

# ----------------------- CONFIGURACIÓN -----------------------
EXCEL_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "binance_futures_gainers.xlsx")
TOP_N = 15                     # cuántos contratos "ganadores" guardar cada corrida
MIN_VOLUME_USDT = 1_000_000    # filtro para evitar contratos ilíquidos/ruido
QUOTE_ASSET = "USDT"           # solo pares contra USDT (mercado USDⓈ-M)
MERCADO = "futuros"            # "futuros" o "spot"
DIRECCION = "ascendente"       # "ascendente" (ganadoras) o "descendente" (perdedoras)

# --- Detección de racha sostenida (para alertas y el ícono 🔥 del dashboard) ---
VENTANA_RACHA = 3              # cuántas corridas seguidas debe subir (o bajar) para calificar
UMBRAL_RACHA = 8.0             # cambio mínimo acumulado en esa ventana, en puntos porcentuales

# --- Retención de historial: evita que el Excel crezca sin límite si el
# script corre 24/7 de forma indefinida (clave para correr en la nube) ---
RETENCION_DIAS_HISTORIAL = 30  # cuántos días de historial conservar en "Datos" (Excel completo)
VENTANA_DASHBOARD_DIAS = 4     # cuántos días mostrar en el gráfico/análisis del dashboard web
                                # (mucho menor que la retención, para que el gráfico no se sature
                                # ni se ponga lento con meses de monedas acumuladas)
MAX_PUNTOS_GRAFICO = 200       # máximo de puntos por línea en el gráfico (reduce puntos si hay
                                # más historial denso del que el navegador puede dibujar fluido)

# --- Alertas por Telegram cuando se detecta una racha sostenida ---
TELEGRAM_BOT_TOKEN = "8831860371:AAE8scHsUzP35JjAJKleuG22y-2a01jQ3iM"
TELEGRAM_CHAT_ID = "1564425184"

# --- Publicación automática en GitHub Pages (dashboard accesible online) ---
PUBLICAR_EN_GITHUB = True
REPO_GITHUB_CARPETA = r"C:\BinanceTracker\github-repo\binance-dashboard"  # carpeta del "git clone" (solo se usa en modo local)
REPO_GITHUB_NOMBRE_ARCHIVO = "gainers.html"  # nombre del archivo dentro del repo (URL pública)


# --- Umbral para resaltar un funding rate "extremo" en el análisis (en %) ---
UMBRAL_FUNDING_ALERTA = 0.05

# --- Indicadores técnicos (RSI, medias móviles, soporte/resistencia) ---
# Calculados desde velas públicas de Binance, no de TradingView (no tienen API pública).
KLINES_INTERVALO = "15m"       # tamaño de vela para los indicadores
KLINES_CANTIDAD = 100          # ~25 horas de velas de 15 min
RSI_PERIODO = 14
SMA_CORTA = 20
SMA_LARGA = 50
UMBRAL_CERCA_NIVEL = 1.0       # % de distancia para considerar "muy cerca" de soporte/resistencia
# ---------------------------------------------------------------


def obtener_klines(symbol):
    """Consulta velas públicas de Binance para un símbolo y devuelve
    (cierres, máximos, mínimos, volúmenes) como listas, o None si falla.
    Esta es la fuente de datos para RSI, medias móviles y soporte/resistencia
    — no requiere ni usa TradingView, que no tiene API pública para esto."""
    base = "https://fapi.binance.com/fapi/v1/klines" if MERCADO == "futuros" else "https://api.binance.com/api/v3/klines"
    try:
        resp = requests.get(
            base, params={"symbol": symbol, "interval": KLINES_INTERVALO, "limit": KLINES_CANTIDAD}, timeout=10
        )
        resp.raise_for_status()
        data = resp.json()
        cierres = [float(k[4]) for k in data]
        maximos = [float(k[2]) for k in data]
        minimos = [float(k[3]) for k in data]
        volumenes = [float(k[5]) for k in data]
        return cierres, maximos, minimos, volumenes
    except Exception:
        return None


def calcular_rsi(cierres, periodo=RSI_PERIODO):
    """RSI clásico (método de Wilder). Por encima de 70 = sobrecompra;
    por debajo de 30 = sobreventa."""
    if len(cierres) < periodo + 1:
        return None
    deltas = [cierres[i] - cierres[i - 1] for i in range(1, len(cierres))]
    ganancias = [d if d > 0 else 0 for d in deltas]
    perdidas = [-d if d < 0 else 0 for d in deltas]
    avg_gain = sum(ganancias[:periodo]) / periodo
    avg_loss = sum(perdidas[:periodo]) / periodo
    for i in range(periodo, len(deltas)):
        avg_gain = (avg_gain * (periodo - 1) + ganancias[i]) / periodo
        avg_loss = (avg_loss * (periodo - 1) + perdidas[i]) / periodo
    if avg_loss == 0:
        return 100.0
    rs = avg_gain / avg_loss
    return round(100 - (100 / (1 + rs)), 1)


def sma(valores, periodo):
    """Media móvil simple de los últimos N valores."""
    if len(valores) < periodo:
        return None
    return sum(valores[-periodo:]) / periodo


def obtener_order_book(symbol, profundidad=20):
    """Consulta el libro de órdenes público de Binance (bids/asks) y devuelve
    (bids, asks) como listas de [precio, cantidad], o None si falla."""
    base = "https://fapi.binance.com/fapi/v1/depth" if MERCADO == "futuros" else "https://api.binance.com/api/v3/depth"
    try:
        resp = requests.get(base, params={"symbol": symbol, "limit": profundidad}, timeout=10)
        resp.raise_for_status()
        data = resp.json()
        bids = [[float(p), float(q)] for p, q in data.get("bids", [])]
        asks = [[float(p), float(q)] for p, q in data.get("asks", [])]
        return bids, asks
    except Exception:
        return None


def calcular_metricas_orderbook(bids, asks):
    """A partir del libro de órdenes calcula:
    - spread_pct: qué tan separados están el mejor precio de compra y de venta
      (más ajustado = mercado más líquido/eficiente)
    - imbalance_pct: de -100 a +100. Positivo = hay más volumen esperando
      COMPRAR (presión alcista); negativo = más volumen esperando VENDER
      (presión bajista). Se calcula sobre los primeros niveles del libro."""
    if not bids or not asks:
        return None, None
    mejor_bid = bids[0][0]
    mejor_ask = asks[0][0]
    spread_pct = (mejor_ask - mejor_bid) / mejor_bid * 100 if mejor_bid else None

    vol_bids = sum(q for _, q in bids)
    vol_asks = sum(q for _, q in asks)
    total = vol_bids + vol_asks
    imbalance_pct = ((vol_bids - vol_asks) / total * 100) if total > 0 else None

    return spread_pct, imbalance_pct


def formatear_precio(valor):
    """Formatea un precio con suficientes decimales según su magnitud —
    una moneda de $65,000 y una de $0.00003 necesitan escalas distintas."""
    if valor is None:
        return "—"
    if valor >= 100:
        return f"${valor:,.2f}"
    elif valor >= 1:
        return f"${valor:,.4f}"
    else:
        return f"${valor:.8f}"


def reducir_puntos(lista, maximo):
    """Reduce una lista de timestamps a un máximo de puntos, tomando uno cada
    N (sin perder el más reciente). Esto es solo para lo que se DIBUJA en el
    gráfico — no afecta ningún cálculo del análisis, que sigue usando los
    datos completos. Necesario porque meses de historial denso (por ejemplo
    de cuando corría cada 5 minutos) pueden tener miles de puntos por línea,
    lo cual pone lento al navegador al interactuar con el gráfico."""
    if len(lista) <= maximo:
        return lista
    paso = (len(lista) + maximo - 1) // maximo  # equivalente a ceil(len/maximo)
    reducida = lista[::paso]
    if reducida[-1] != lista[-1]:
        reducida.append(lista[-1])
    return reducida


def obtener_funding_rates():
    """Consulta la API pública de Binance Futures y devuelve el funding rate
    actual de TODOS los contratos perpetuos en una sola llamada: {symbol: %}.
    Un funding rate muy positivo indica que el mercado está cargado de
    posiciones largas (riesgo de long squeeze); muy negativo, lo contrario.
    Solo aplica al mercado de futuros — no existe en spot."""
    if MERCADO != "futuros":
        return {}
    try:
        resp = requests.get("https://fapi.binance.com/fapi/v1/premiumIndex", timeout=15)
        resp.raise_for_status()
        data = resp.json()
        return {
            item["symbol"]: float(item["lastFundingRate"]) * 100
            for item in data if "lastFundingRate" in item
        }
    except Exception as e:
        print(f"Aviso: no se pudo obtener el funding rate ({e}). Se omite esa columna esta corrida.")
        return {}


def obtener_top_ganadoras():
    """Consulta la API pública de Binance Futures y devuelve el top N por % de cambio en 24h."""
    if MERCADO == "futuros":
        url = "https://fapi.binance.com/fapi/v1/ticker/24hr"
    else:
        url = "https://api.binance.com/api/v3/ticker/24hr"
    resp = requests.get(url, timeout=15)
    resp.raise_for_status()
    data = resp.json()

    candidatos = []
    for item in data:
        symbol = item["symbol"]
        if not symbol.endswith(QUOTE_ASSET):
            continue
        try:
            cambio_pct = float(item["priceChangePercent"])
            volumen = float(item["quoteVolume"])
            precio = float(item["lastPrice"])
        except (ValueError, KeyError):
            continue
        if volumen < MIN_VOLUME_USDT:
            continue
        candidatos.append({
            "symbol": symbol,
            "cambio_pct": cambio_pct,
            "precio": precio,
            "volumen": volumen,
        })

    candidatos.sort(key=lambda x: x["cambio_pct"], reverse=True)
    return candidatos[:TOP_N]


def limpiar_historial_antiguo(wb, ws):
    """Elimina de la hoja 'Datos' las filas más viejas que RETENCION_DIAS_HISTORIAL.
    Sin esto, si el script corre 24/7 de forma indefinida en la nube, el Excel
    crecería sin límite para siempre — esto lo mantiene acotado. Devuelve la
    hoja actualizada (nueva si tuvo que recrearse)."""
    limite = datetime.now() - timedelta(days=RETENCION_DIAS_HISTORIAL)
    filas = list(ws.iter_rows(min_row=2, values_only=True))
    if not filas:
        return ws
    try:
        filas_nuevas = [f for f in filas if datetime.strptime(f[0], "%Y-%m-%d %H:%M:%S") >= limite]
    except Exception:
        return ws  # si algo no calza con el formato de fecha, no arriesgar el archivo

    if len(filas_nuevas) == len(filas):
        return ws  # nada que recortar todavía

    encabezado = [c.value for c in ws[1]]
    del wb["Datos"]
    ws_nueva = wb.create_sheet("Datos", 0)
    ws_nueva.append(encabezado)
    for cell in ws_nueva[1]:
        cell.font = cell.font.copy(bold=True)
    for fila in filas_nuevas:
        ws_nueva.append(list(fila))
    print(f"Historial recortado: se eliminaron {len(filas) - len(filas_nuevas)} filas de más de {RETENCION_DIAS_HISTORIAL} días.")
    return ws_nueva


def guardar_en_excel(top_ganadoras):
    """Añade una fila por cada moneda del top, con timestamp, y actualiza la gráfica."""
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    if os.path.exists(EXCEL_FILE):
        try:
            wb = load_workbook(EXCEL_FILE)
            ws = wb["Datos"]
        except Exception as e:
            print(f"[{timestamp}] No se pudo abrir el Excel existente ({e}). "
                  f"Puede estar dañado o abierto en otro programa. Se omite esta corrida.")
            return
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Datos"
        ws.append(["Fecha/Hora", "Symbol", "% Cambio 24h", "Precio", "Volumen USDT", "Funding Rate %"])
        for cell in ws[1]:
            cell.font = cell.font.copy(bold=True)

    # Una sola consulta de funding rate por corrida, reutilizada también en el dashboard
    funding_dict = obtener_funding_rates()

    for moneda in top_ganadoras:
        funding = funding_dict.get(moneda["symbol"])
        ws.append([
            timestamp,
            moneda["symbol"],
            round(moneda["cambio_pct"], 2),
            moneda["precio"],
            round(moneda["volumen"], 2),
            round(funding, 4) if funding is not None else None,
        ])

    # Recorta filas viejas antes de reconstruir Pivot/Aceleración/etc, para que
    # todo el dashboard refleje siempre un historial acotado (clave en la nube)
    ws = limpiar_historial_antiguo(wb, ws)

    # Ajustar ancho de columnas
    anchos = [20, 14, 14, 16, 16, 15]
    for i, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[chr(64 + i)].width = ancho

    actualizar_grafica(wb, ws)

    # Guardado atómico: primero a un archivo temporal, y solo si se completa
    # bien, se reemplaza el archivo real. Así, si la laptop se apaga, se
    # queda sin batería, o se reinicia justo durante el guardado, el archivo
    # bueno anterior queda intacto en vez de corromperse a medias.
    archivo_temporal = EXCEL_FILE + ".tmp"
    try:
        wb.save(archivo_temporal)
        os.replace(archivo_temporal, EXCEL_FILE)
    except Exception as e:
        print(f"[{timestamp}] ERROR al guardar, se conserva el archivo anterior sin cambios: {e}")
        if os.path.exists(archivo_temporal):
            os.remove(archivo_temporal)
        return

    generar_dashboard_html(wb, ws)
    print(f"[{timestamp}] Guardadas {len(top_ganadoras)} monedas en {EXCEL_FILE}")


def leer_tabla_agrupada(ws):
    """Lee la hoja 'Datos' y la agrupa en: {timestamp: {symbol: % cambio}},
    la lista de símbolos vistos (en orden de aparición), los timestamps
    ordenados, el precio/volumen/funding más reciente de cada símbolo, y el
    historial de funding rate {timestamp: {symbol: funding}} (para calcular
    su tendencia). Función compartida por Pivot, el dashboard HTML y las
    alertas, para que todos usen exactamente los mismos datos.

    Compatible con archivos viejos sin columna de Funding Rate: en esas filas
    el valor simplemente llega como None."""
    filas = list(ws.iter_rows(min_row=2, values_only=True))
    tabla = {}
    tabla_funding = {}
    simbolos_vistos = []
    ultimo_info = {}
    for fila in filas:
        fecha, symbol, cambio_pct, precio, volumen = fila[:5]
        funding = fila[5] if len(fila) > 5 else None
        if fecha not in tabla:
            tabla[fecha] = {}
            tabla_funding[fecha] = {}
        tabla[fecha][symbol] = cambio_pct
        tabla_funding[fecha][symbol] = funding
        if symbol not in simbolos_vistos:
            simbolos_vistos.append(symbol)
        ultimo_info[symbol] = {"precio": precio, "volumen": volumen, "funding": funding}
    timestamps_ordenados = sorted(tabla.keys())
    return tabla, simbolos_vistos, timestamps_ordenados, ultimo_info, tabla_funding


def calcular_rachas(tabla, timestamps_ordenados, simbolos_vistos):
    """Devuelve la lista de monedas cuya racha de las últimas VENTANA_RACHA
    corridas es consistentemente ascendente (o descendente, según DIRECCION)
    Y supera el umbral configurado. Evita alertas por un solo salto de ruido:
    solo califica si el movimiento fue sostenido corrida tras corrida."""
    calificados = []
    if len(timestamps_ordenados) < VENTANA_RACHA:
        return calificados

    ventana_ts = timestamps_ordenados[-VENTANA_RACHA:]
    for symbol in simbolos_vistos:
        valores = [tabla[ts].get(symbol) for ts in ventana_ts]
        if any(v is None for v in valores):
            continue  # no apareció en el top en alguna de esas corridas, no se puede evaluar

        if DIRECCION == "ascendente":
            monotona = all(valores[i] < valores[i + 1] for i in range(len(valores) - 1))
            delta_total = round(valores[-1] - valores[0], 2)
            califica = monotona and delta_total >= UMBRAL_RACHA
        else:
            monotona = all(valores[i] > valores[i + 1] for i in range(len(valores) - 1))
            delta_total = round(valores[-1] - valores[0], 2)
            califica = monotona and delta_total <= -UMBRAL_RACHA

        if califica:
            calificados.append({
                "symbol": symbol,
                "valores": valores,
                "delta_total": delta_total,
            })

    return calificados


def enviar_notificacion_windows(calificados):
    """Muestra una notificación emergente de Windows si hay monedas en racha.
    Requiere 'plyer' (pip install plyer); si no está instalado, solo lo avisa
    por consola y sigue sin interrumpir el resto del script."""
    try:
        from plyer import notification
    except ImportError:
        print("Aviso: instala 'plyer' (pip install plyer) para recibir notificaciones "
              "de Windows con las alertas de racha.")
        return

    simbolos_txt = ", ".join(item["symbol"] for item in calificados[:5])
    verbo = "subiendo" if DIRECCION == "ascendente" else "cayendo"
    try:
        notification.notify(
            title="🔥 Binance Tracker - Racha detectada",
            message=f"{simbolos_txt} llevan {VENTANA_RACHA} corridas seguidas {verbo} fuerte.",
            timeout=15,
        )
    except Exception as e:
        print(f"No se pudo mostrar la notificación de Windows: {e}")


def enviar_notificacion_telegram(calificados):
    """Manda un mensaje al chat de Telegram configurado (TELEGRAM_BOT_TOKEN /
    TELEGRAM_CHAT_ID) si hay monedas en racha. No interrumpe el script si
    falla (ej. sin internet, token inválido): solo lo avisa por consola."""
    if not TELEGRAM_BOT_TOKEN or not TELEGRAM_CHAT_ID:
        return

    verbo = "subiendo" if DIRECCION == "ascendente" else "cayendo"
    lineas = [f"🔥 *Racha detectada* ({verbo} fuerte, {VENTANA_RACHA} corridas seguidas):", ""]
    for item in calificados:
        lineas.append(f"• *{item['symbol']}*: {item['delta_total']:+.2f} p.p.")
    mensaje = "\n".join(lineas)

    url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage"
    try:
        resp = requests.post(url, json={
            "chat_id": TELEGRAM_CHAT_ID,
            "text": mensaje,
            "parse_mode": "Markdown",
        }, timeout=15)
        if not resp.ok:
            print(f"Aviso: Telegram respondió con error al enviar la alerta: {resp.text}")
    except Exception as e:
        print(f"Aviso: no se pudo enviar la alerta a Telegram: {e}")


def procesar_alertas(wb, tabla, timestamps_ordenados, simbolos_vistos):
    """Detecta rachas sostenidas, las registra en la hoja 'Alertas' (que se
    conserva entre corridas, a diferencia de Pivot/Aceleración/Gráfica/Selector
    que se reconstruyen desde cero cada vez), y dispara la notificación."""
    calificados = calcular_rachas(tabla, timestamps_ordenados, simbolos_vistos)

    if "Alertas" not in wb.sheetnames:
        hoja_alertas = wb.create_sheet("Alertas")
        hoja_alertas.append(["Fecha/Hora detección", "Symbol", "Valores en la racha (%)", "Cambio total (p.p.)"])
        for cell in hoja_alertas[1]:
            cell.font = Font(bold=True)
        hoja_alertas.column_dimensions["A"].width = 20
        hoja_alertas.column_dimensions["B"].width = 14
        hoja_alertas.column_dimensions["C"].width = 32
        hoja_alertas.column_dimensions["D"].width = 20
    else:
        hoja_alertas = wb["Alertas"]

    if not calificados:
        return

    timestamp_actual = timestamps_ordenados[-1]
    for item in calificados:
        valores_str = " → ".join(str(v) for v in item["valores"])
        hoja_alertas.append([timestamp_actual, item["symbol"], valores_str, item["delta_total"]])
        print(f"[{timestamp_actual}] 🔥 RACHA: {item['symbol']} ({valores_str}, "
              f"total {item['delta_total']:+} p.p.)")

    enviar_notificacion_windows(calificados)
    enviar_notificacion_telegram(calificados)


def construir_pivot(wb, ws):
    """Reconstruye una tabla dinámica: filas = timestamps, columnas = monedas,
    valores = % Cambio 24h. Así cada moneda queda en su propia columna y se
    le puede graficar una línea individual."""
    if "Pivot" in wb.sheetnames:
        del wb["Pivot"]
    hoja_pivot = wb.create_sheet("Pivot")

    tabla, simbolos_vistos, timestamps_ordenados, _, _ = leer_tabla_agrupada(ws)
    if not timestamps_ordenados:
        return None

    # Encabezado: Fecha/Hora + una columna por símbolo
    hoja_pivot.append(["Fecha/Hora"] + simbolos_vistos)
    for cell in hoja_pivot[1]:
        cell.font = cell.font.copy(bold=True)

    # Una fila por timestamp, con el % de cada moneda (vacío si no apareció esa corrida)
    for ts in timestamps_ordenados:
        fila = [ts]
        for symbol in simbolos_vistos:
            fila.append(tabla[ts].get(symbol, None))
        hoja_pivot.append(fila)

    hoja_pivot.column_dimensions["A"].width = 20
    for i in range(2, len(simbolos_vistos) + 2):
        hoja_pivot.column_dimensions[get_column_letter(i)].width = 14

    return hoja_pivot, len(timestamps_ordenados), len(simbolos_vistos), tabla, simbolos_vistos, timestamps_ordenados


def construir_aceleracion(wb, hoja_pivot, num_filas, num_simbolos):
    """Calcula, para cada moneda, cuánto subió o bajó su % de cambio 24h entre
    la corrida anterior y la más reciente (aceleración de corto plazo).
    Ordena de mayor a menor aceleración para que las que están 'despegando'
    aparezcan primero, y resalta con color las más fuertes."""
    if "Aceleracion" in wb.sheetnames:
        del wb["Aceleracion"]
    hoja_acel = wb.create_sheet("Aceleracion", 0)  # la deja como primera hoja para verla de un vistazo

    if num_filas < 2:
        hoja_acel.append(["Se necesitan al menos 2 corridas para calcular aceleración."])
        return

    simbolos = [hoja_pivot.cell(row=1, column=c).value for c in range(2, num_simbolos + 2)]

    filas_calculadas = []
    for idx, symbol in enumerate(simbolos):
        col = idx + 2
        # Tomar los últimos dos valores NO vacíos de esa columna (recorriendo de abajo hacia arriba)
        valores_recientes = []
        for row in range(num_filas + 1, 1, -1):
            val = hoja_pivot.cell(row=row, column=col).value
            if val is not None:
                valores_recientes.append(val)
            if len(valores_recientes) == 2:
                break
        if len(valores_recientes) < 2:
            continue  # esta moneda solo tiene una aparición, no se puede medir aceleración aún
        actual, anterior = valores_recientes
        delta = round(actual - anterior, 2)
        filas_calculadas.append((symbol, anterior, actual, delta))

    filas_calculadas.sort(key=lambda x: x[3], reverse=True)

    encabezado = ["Symbol", "% Anterior", "% Actual", "Aceleración (p.p. desde última corrida)"]
    hoja_acel.append(encabezado)
    for cell in hoja_acel[1]:
        cell.font = Font(bold=True)

    verde = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    rojo = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for symbol, anterior, actual, delta in filas_calculadas:
        hoja_acel.append([symbol, anterior, actual, delta])
        celda_delta = hoja_acel.cell(row=hoja_acel.max_row, column=4)
        if delta > 0:
            celda_delta.fill = verde
        elif delta < 0:
            celda_delta.fill = rojo

    anchos = [16, 14, 14, 30]
    for i, ancho in enumerate(anchos, start=1):
        hoja_acel.column_dimensions[chr(64 + i)].width = ancho


def construir_selector(wb, num_filas, num_simbolos):
    """Crea una hoja 'Selector' con un menú desplegable (dropdown) para elegir
    una moneda. El gráfico y los datos se actualizan solos usando fórmulas
    nativas de Excel (INDEX/MATCH) — no requiere volver a correr el script,
    solo cambiar el desplegable y Excel recalcula automáticamente."""
    if "Selector" in wb.sheetnames:
        del wb["Selector"]
    hoja_sel = wb.create_sheet("Selector", 1)  # justo después de "Aceleracion"

    if num_filas < 1 or num_simbolos < 1:
        hoja_sel["A1"] = "Aún no hay suficientes datos para el selector."
        return

    ultima_col_pivot = get_column_letter(num_simbolos + 1)  # Pivot: col B en adelante son símbolos

    hoja_sel["A1"] = "Selecciona una moneda:"
    hoja_sel["A1"].font = Font(bold=True)
    hoja_sel["B1"] = f"=Pivot!B1"  # arranca mostrando el primer símbolo por defecto

    dv = DataValidation(
        type="list",
        formula1=f"=Pivot!$B$1:${ultima_col_pivot}$1",
        allow_blank=False,
    )
    hoja_sel.add_data_validation(dv)
    dv.add(hoja_sel["B1"])
    hoja_sel["B1"].fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    hoja_sel["A3"] = "Fecha/Hora"
    hoja_sel["B3"] = "% Cambio (moneda seleccionada)"
    for cell in hoja_sel["A3:B3"][0]:
        cell.font = Font(bold=True)

    for i in range(num_filas):
        fila_excel = i + 4  # empieza en la fila 4
        fila_pivot = i + 2  # en Pivot, los datos empiezan en la fila 2
        hoja_sel.cell(row=fila_excel, column=1, value=f"=Pivot!A{fila_pivot}")
        formula_valor = (
            f'=IFERROR(INDEX(Pivot!$B$2:${ultima_col_pivot}${num_filas + 1},{i + 1},'
            f'MATCH($B$1,Pivot!$B$1:${ultima_col_pivot}$1,0)),"")'
        )
        hoja_sel.cell(row=fila_excel, column=2, value=formula_valor)

    hoja_sel.column_dimensions["A"].width = 20
    hoja_sel.column_dimensions["B"].width = 26

    chart = LineChart()
    chart.title = "Evolución de la moneda seleccionada (cambia el desplegable en B1)"
    chart.x_axis.title = "Fecha/Hora"
    chart.y_axis.title = "% Cambio 24h"

    datos = Reference(hoja_sel, min_col=2, min_row=3, max_row=num_filas + 3)
    categorias = Reference(hoja_sel, min_col=1, min_row=4, max_row=num_filas + 3)
    chart.add_data(datos, titles_from_data=True)
    chart.set_categories(categorias)
    chart.width = 26
    chart.height = 14

    hoja_sel.add_chart(chart, "D3")


def actualizar_grafica(wb, ws):
    """Construye la tabla Pivot (una columna por moneda), la hoja de Aceleración,
    el selector interactivo, procesa las alertas de racha, y una gráfica de
    líneas con una serie por símbolo."""
    if "Grafica" in wb.sheetnames:
        del wb["Grafica"]

    resultado = construir_pivot(wb, ws)
    if resultado is None:
        return
    hoja_pivot, num_filas, num_simbolos, tabla, simbolos_vistos, timestamps_ordenados = resultado

    construir_aceleracion(wb, hoja_pivot, num_filas, num_simbolos)
    construir_selector(wb, num_filas, num_simbolos)
    procesar_alertas(wb, tabla, timestamps_ordenados, simbolos_vistos)

    hoja_grafica = wb.create_sheet("Grafica")

    chart = LineChart()
    chart.title = "% Cambio 24h por moneda a lo largo del tiempo"
    chart.x_axis.title = "Fecha/Hora"
    chart.y_axis.title = "% Cambio 24h"

    max_row = num_filas + 1  # +1 por el encabezado
    max_col = num_simbolos + 1  # +1 por la columna de Fecha/Hora

    datos = Reference(hoja_pivot, min_col=2, max_col=max_col, min_row=1, max_row=max_row)
    categorias = Reference(hoja_pivot, min_col=1, min_row=2, max_row=max_row)
    chart.add_data(datos, titles_from_data=True)
    chart.set_categories(categorias)
    chart.width = 30
    chart.height = 15

    hoja_grafica.add_chart(chart, "A1")


def generar_dashboard_html(wb, ws):
    """Genera un archivo HTML interactivo (Chart.js) con una línea por moneda,
    más una sección de alertas arriba del gráfico (activas ahora + historial
    reciente), para poder revisar todo desde el celular sin abrir el Excel.
    Al hacer clic en el nombre de una moneda en la leyenda, esa línea se
    oculta/muestra sola, permitiendo aislar visualmente su evolución.
    Las monedas en racha sostenida (ver calcular_rachas) aparecen con 🔥
    delante de su nombre en la leyenda.

    Para que el gráfico no se sature con monedas viejas que ya salieron
    del top, por defecto solo se muestran (visibles) las que están en el
    top de la corrida más reciente o en racha activa. Las demás quedan
    ocultas de inicio, pero siguen en la leyenda por si quieres revisarlas
    manualmente con un clic."""
    tabla, simbolos_vistos, timestamps_ordenados, ultimo_info, tabla_funding = leer_tabla_agrupada(ws)
    if not timestamps_ordenados:
        return

    # Limitar el dashboard a los últimos VENTANA_DASHBOARD_DIAS días. El Excel
    # conserva hasta RETENCION_DIAS_HISTORIAL días completos, pero mostrar todo
    # eso en el gráfico web (con meses de monedas acumuladas) lo satura y pone
    # lento al navegador al buscar/pasar el mouse. Esto no afecta el Excel.
    limite_dashboard = datetime.now() - timedelta(days=VENTANA_DASHBOARD_DIAS)
    timestamps_ordenados = [
        ts for ts in timestamps_ordenados
        if datetime.strptime(ts, "%Y-%m-%d %H:%M:%S") >= limite_dashboard
    ]
    if not timestamps_ordenados:
        return
    simbolos_vistos = [
        s for s in simbolos_vistos
        if any(tabla[ts].get(s) is not None for ts in timestamps_ordenados)
    ]

    calificados = calcular_rachas(tabla, timestamps_ordenados, simbolos_vistos)
    simbolos_en_racha = {item["symbol"] for item in calificados}

    # --- Sección de análisis: combina racha, funding rate (y su tendencia),
    # volumen, volatilidad reciente y extremo del período en un veredicto
    # simple y explicado, para evaluar una posible entrada. Es contexto para
    # apoyar tu criterio, NO una señal automática de compra/venta. ---
    VOLUMEN_SALUDABLE = 5_000_000  # por encima de esto, se considera buena liquidez (no solo pasar el filtro mínimo)
    VOLUMEN_ALTO = 50_000_000      # por encima de esto, suma punto extra a favor

    ultimo_timestamp_para_analisis = timestamps_ordenados[-1]
    simbolos_top_para_analisis = [
        s for s in simbolos_vistos if tabla[ultimo_timestamp_para_analisis].get(s) is not None
    ]

    filas_analisis = []
    for symbol in simbolos_top_para_analisis:
        pct_actual = tabla[ultimo_timestamp_para_analisis].get(symbol)
        info = ultimo_info.get(symbol, {})
        volumen = info.get("volumen")
        funding = info.get("funding")
        en_racha = symbol in simbolos_en_racha

        # Extremo del período
        valores_historicos = [tabla[ts].get(symbol) for ts in timestamps_ordenados if tabla[ts].get(symbol) is not None]
        if DIRECCION == "ascendente":
            en_extremo = bool(valores_historicos) and pct_actual == max(valores_historicos)
            texto_extremo = "🆙 Máximo del período" if en_extremo else ""
        else:
            en_extremo = bool(valores_historicos) and pct_actual == min(valores_historicos)
            texto_extremo = "🔻 Mínimo del período" if en_extremo else ""

        # Volatilidad reciente: rango (máx-mín) de las últimas corridas disponibles (hasta 6)
        ventana_vol = valores_historicos[-6:] if len(valores_historicos) >= 2 else valores_historicos
        rango_volatilidad = round(max(ventana_vol) - min(ventana_vol), 2) if len(ventana_vol) >= 2 else None

        # --- Indicadores técnicos desde velas de Binance: RSI, medias móviles, soporte/resistencia ---
        klines = obtener_klines(symbol)
        rsi_val = tendencia_ma = texto_nivel = dist_nivel_pct = None
        precio_soporte = precio_resistencia = None
        clase_rsi = ""
        if klines:
            cierres_k, maximos_k, minimos_k, volumenes_k = klines
            rsi_val = calcular_rsi(cierres_k)
            sma_corta_val = sma(cierres_k, SMA_CORTA)
            sma_larga_val = sma(cierres_k, SMA_LARGA)
            precio_k = cierres_k[-1] if cierres_k else None

            if precio_k is not None and sma_corta_val is not None and sma_larga_val is not None:
                if precio_k > sma_corta_val > sma_larga_val:
                    tendencia_ma = "📈 Alcista"
                elif precio_k < sma_corta_val < sma_larga_val:
                    tendencia_ma = "📉 Bajista"
                else:
                    tendencia_ma = "➡️ Mixta"

            if precio_k and maximos_k and minimos_k:
                precio_resistencia = max(maximos_k)
                precio_soporte = min(minimos_k)
                if DIRECCION == "ascendente":
                    dist_nivel_pct = (precio_resistencia - precio_k) / precio_k * 100
                    texto_nivel = f"Resistencia a {dist_nivel_pct:.2f}%"
                else:
                    dist_nivel_pct = (precio_k - precio_soporte) / precio_k * 100
                    texto_nivel = f"Soporte a {dist_nivel_pct:.2f}%"

            if rsi_val is not None:
                rsi_extremo = (DIRECCION == "ascendente" and rsi_val >= 70) or (DIRECCION == "descendente" and rsi_val <= 30)
                clase_rsi = "delta-neg" if rsi_extremo else ""

        # --- Libro de órdenes: spread y desequilibrio compra/venta ---
        orderbook = obtener_order_book(symbol)
        spread_pct = imbalance_pct = None
        if orderbook:
            bids, asks = orderbook
            spread_pct, imbalance_pct = calcular_metricas_orderbook(bids, asks)

        # Funding rate: valor + tendencia (comparando contra unas corridas atrás)
        funding_extremo = funding is not None and abs(funding) >= UMBRAL_FUNDING_ALERTA
        fundings_recientes = [
            tabla_funding.get(ts, {}).get(symbol) for ts in timestamps_ordenados[-VENTANA_RACHA:]
        ]
        fundings_recientes = [f for f in fundings_recientes if f is not None]
        if len(fundings_recientes) >= 2:
            delta_funding = fundings_recientes[-1] - fundings_recientes[0]
            if abs(delta_funding) < 0.005:
                tendencia_funding = "→ estable"
            elif delta_funding > 0:
                tendencia_funding = "↑ subiendo"
            else:
                tendencia_funding = "↓ bajando"
        else:
            tendencia_funding = "—"

        # --- Veredicto compuesto (explicado, no una caja negra) ---
        puntaje = 0
        motivos = []
        if en_racha:
            puntaje += 2
            motivos.append("racha sostenida a favor")
        if funding_extremo:
            # funding extremo en la MISMA dirección del movimiento = mercado ya muy cargado
            mismo_sentido = (funding > 0 and DIRECCION == "ascendente") or (funding < 0 and DIRECCION == "descendente")
            if mismo_sentido:
                puntaje -= 2
                motivos.append("funding ya muy cargado en la misma dirección (riesgo de squeeze)")
        if en_extremo:
            puntaje -= 1
            motivos.append("en el extremo de su rango reciente (posible agotamiento)")
        if tendencia_ma == "📈 Alcista" and DIRECCION == "ascendente":
            puntaje += 1
            motivos.append("estructura de medias móviles alcista (20 y 50 alineadas)")
        elif tendencia_ma == "📉 Bajista" and DIRECCION == "descendente":
            puntaje += 1
            motivos.append("estructura de medias móviles bajista (a favor del corto)")
        if rsi_val is not None:
            if DIRECCION == "ascendente" and rsi_val >= 70:
                puntaje -= 1
                motivos.append(f"RSI en sobrecompra ({rsi_val})")
            elif DIRECCION == "descendente" and rsi_val <= 30:
                puntaje -= 1
                motivos.append(f"RSI en sobreventa ({rsi_val})")
        if dist_nivel_pct is not None and dist_nivel_pct <= UMBRAL_CERCA_NIVEL:
            puntaje -= 1
            nivel_nombre = "resistencia" if DIRECCION == "ascendente" else "soporte"
            motivos.append(f"muy cerca de un nivel clave de {nivel_nombre} ({dist_nivel_pct:.2f}%)")
        if imbalance_pct is not None:
            if DIRECCION == "ascendente" and imbalance_pct <= -20:
                puntaje -= 1
                motivos.append(f"libro de órdenes con más presión de venta que de compra ({imbalance_pct:+.1f}%), a pesar de la subida")
            elif DIRECCION == "ascendente" and imbalance_pct >= 20:
                puntaje += 1
                motivos.append(f"libro de órdenes confirma presión de compra ({imbalance_pct:+.1f}%)")
            elif DIRECCION == "descendente" and imbalance_pct >= 20:
                puntaje -= 1
                motivos.append(f"libro de órdenes con más presión de compra que de venta ({imbalance_pct:+.1f}%), a pesar de la caída")
            elif DIRECCION == "descendente" and imbalance_pct <= -20:
                puntaje += 1
                motivos.append(f"libro de órdenes confirma presión de venta ({imbalance_pct:+.1f}%)")
        if volumen is not None:
            if volumen < VOLUMEN_SALUDABLE:
                puntaje -= 1
                motivos.append("volumen relativamente bajo (menos confiable)")
            elif volumen >= VOLUMEN_ALTO:
                puntaje += 1
                motivos.append("volumen alto (movimiento más confiable)")

        if puntaje >= 2:
            veredicto, clase_veredicto = "🟢 Favorable", "veredicto-fav"
        elif puntaje >= 0:
            veredicto, clase_veredicto = "🟡 Vigilar", "veredicto-neu"
        else:
            veredicto, clase_veredicto = "🔴 Precaución", "veredicto-riesgo"

        if funding is not None:
            clase_funding = "delta-neg" if funding_extremo else ""
            funding_txt = f"{funding:+.4f}% ({tendencia_funding})"
        else:
            clase_funding = ""
            funding_txt = "—"

        volumen_txt = f"${volumen:,.0f}" if volumen is not None else "—"
        volatilidad_txt = f"{rango_volatilidad:.2f} p.p." if rango_volatilidad is not None else "—"
        texto_motivos = "; ".join(motivos) if motivos else "sin señales relevantes"
        rsi_txt = str(rsi_val) if rsi_val is not None else "—"
        tendencia_ma_txt = tendencia_ma if tendencia_ma else "—"
        texto_nivel_txt = texto_nivel if texto_nivel else "—"
        soporte_txt = formatear_precio(precio_soporte)
        resistencia_txt = formatear_precio(precio_resistencia)
        spread_txt = f"{spread_pct:.3f}%" if spread_pct is not None else "—"
        if imbalance_pct is not None:
            clase_imbalance = "delta-pos" if imbalance_pct > 0 else "delta-neg"
            imbalance_txt = f"{imbalance_pct:+.1f}% {'compra' if imbalance_pct > 0 else 'venta'}"
        else:
            clase_imbalance = ""
            imbalance_txt = "—"

        filas_analisis.append(
            f"<tr><td>{'🔥 ' if en_racha else ''}{symbol}</td>"
            f"<td>{pct_actual:+.2f}%</td>"
            f"<td>{volumen_txt}</td>"
            f"<td class='{clase_funding}'>{funding_txt}</td>"
            f"<td class='{clase_rsi}'>{rsi_txt}</td>"
            f"<td>{tendencia_ma_txt}</td>"
            f"<td>{texto_nivel_txt}</td>"
            f"<td>{soporte_txt}</td>"
            f"<td>{resistencia_txt}</td>"
            f"<td>{spread_txt}</td>"
            f"<td class='{clase_imbalance}'>{imbalance_txt}</td>"
            f"<td>{volatilidad_txt}</td>"
            f"<td>{texto_extremo}</td>"
            f"<td class='{clase_veredicto}' title='{texto_motivos}'>{veredicto}</td></tr>"
        )

    if filas_analisis:
        html_analisis = f"""
        <p class="nota-analisis"><b>Esto es contexto para apoyar tu criterio — no es una señal automática de compra/venta.</b>
        <br>• <b>Funding Rate</b>: resaltado cuando supera ±{UMBRAL_FUNDING_ALERTA}% — mercado muy cargado de largos o cortos, mayor riesgo de squeeze. La tendencia (↑↓→) compara contra hace {VENTANA_RACHA} corridas.
        <br>• <b>RSI (14, velas de {KLINES_INTERVALO})</b>: resaltado si está en sobrecompra (≥70, para ganadoras) o sobreventa (≤30, para perdedoras) — momentum ya extendido.
        <br>• <b>Tendencia (medias móviles)</b>: compara el precio contra sus promedios de 20 y 50 velas — te dice si la estructura de fondo acompaña el movimiento o no.
        <br>• <b>Nivel clave / Soporte / Resistencia</b>: el % es la distancia al nivel más relevante para la dirección; los precios de Soporte y Resistencia son el mínimo y máximo reales de las últimas ~25h — útiles para ubicar stop loss y objetivos.
        <br>• <b>Spread</b>: separación entre el mejor precio de compra y venta ahora mismo — más ajustado (bajo) = mercado más líquido y eficiente para entrar/salir.
        <br>• <b>Presión Compra/Venta</b>: del libro de órdenes actual — positivo (verde) significa más volumen esperando comprar; negativo (rojo) más volumen esperando vender. Resaltado cuando contradice la dirección del movimiento.
        <br>• <b>Volatilidad reciente</b>: rango entre el % más alto y más bajo de las últimas corridas — más alto = movimiento más "picado".
        <br>• <b>Extremo del período</b>: el % actual es el más alto/bajo de los últimos {VENTANA_DASHBOARD_DIAS} días mostrados en el gráfico.
        <br>• <b>Veredicto</b>: combina todo lo anterior en una lectura simple. Pasa el mouse sobre él para ver el motivo exacto.</p>
        <div class="alertas-caja">
            <table><thead><tr><th>Moneda</th><th>% Actual</th><th>Volumen 24h</th><th>Funding Rate</th><th>RSI</th><th>Tendencia</th><th>Nivel clave</th><th>Soporte</th><th>Resistencia</th><th>Spread</th><th>Presión C/V</th><th>Volatilidad</th><th>Extremo</th><th>Veredicto</th></tr></thead>
            <tbody>{''.join(filas_analisis)}</tbody></table>
        </div>"""
    else:
        html_analisis = """<div class="alertas-caja"><p class="sin-alertas">Sin datos suficientes todavía.</p></div>"""

    # --- Sección de alertas: activas ahora mismo (recién calculadas arriba) ---
    if calificados:
        filas_activas = "".join(
            f"<tr><td>🔥 {item['symbol']}</td>"
            f"<td>{' → '.join(str(v) for v in item['valores'])}</td>"
            f"<td class='delta-{'pos' if item['delta_total'] >= 0 else 'neg'}'>{item['delta_total']:+.2f} p.p.</td></tr>"
            for item in calificados
        )
        html_alertas_activas = f"""
        <div class="alertas-caja alertas-activas">
            <h3>🔥 En racha ahora mismo</h3>
            <table><thead><tr><th>Moneda</th><th>Valores recientes (%)</th><th>Cambio en la racha</th></tr></thead>
            <tbody>{filas_activas}</tbody></table>
        </div>"""
    else:
        html_alertas_activas = """
        <div class="alertas-caja alertas-activas">
            <h3>🔥 En racha ahora mismo</h3>
            <p class="sin-alertas">Ninguna moneda califica en este momento.</p>
        </div>"""

    # --- Historial reciente, leído de la hoja "Alertas" (se acumula entre corridas) ---
    html_alertas_historial = ""
    if "Alertas" in wb.sheetnames:
        hoja_alertas = wb["Alertas"]
        filas_hist = list(hoja_alertas.iter_rows(min_row=2, values_only=True))
        ultimas = list(reversed(filas_hist))[:15]  # las 15 más recientes primero
        if ultimas:
            filas_html = "".join(
                f"<tr><td>{fecha}</td><td>{symbol}</td><td>{valores}</td>"
                f"<td class='delta-{'pos' if delta >= 0 else 'neg'}'>{delta:+.2f} p.p.</td></tr>"
                for fecha, symbol, valores, delta in ultimas
            )
            html_alertas_historial = f"""
            <div class="alertas-caja">
                <h3>Historial reciente de alertas</h3>
                <table><thead><tr><th>Fecha/Hora</th><th>Moneda</th><th>Valores</th><th>Cambio</th></tr></thead>
                <tbody>{filas_html}</tbody></table>
            </div>"""

    ultimo_timestamp = timestamps_ordenados[-1]
    simbolos_top_actual = {
        symbol for symbol in simbolos_vistos
        if tabla[ultimo_timestamp].get(symbol) is not None
    }
    simbolos_relevantes = simbolos_top_actual | simbolos_en_racha

    # Puntos reducidos solo para dibujar el gráfico (el análisis y los chips
    # siguen usando timestamps_ordenados completo, sin reducir)
    timestamps_grafico = reducir_puntos(timestamps_ordenados, MAX_PUNTOS_GRAFICO)

    datasets_js = []
    chips_html = []
    for i, symbol in enumerate(simbolos_vistos):
        hue = (i * 47) % 360  # separa bien los colores aunque haya muchas monedas
        color = f"hsl({hue}, 70%, 45%)"
        valores = [tabla[ts].get(symbol, "null") for ts in timestamps_grafico]
        valores_js = "[" + ",".join("null" if v == "null" else str(v) for v in valores) + "]"
        etiqueta = f"🔥 {symbol}" if symbol in simbolos_en_racha else symbol
        oculto_inicial = "true" if symbol not in simbolos_relevantes else "false"
        datasets_js.append(
            f"""{{
                label: "{etiqueta}",
                data: {valores_js},
                borderColor: "{color}",
                backgroundColor: "{color}",
                spanGaps: true,
                tension: 0.2,
                pointRadius: 2,
                pointHitRadius: 8,
                borderWidth: 2,
                hidden: {oculto_inicial}
            }}"""
        )
        clase_oculto = "chip-oculto" if oculto_inicial == "true" else ""
        chips_html.append(
            f'<span class="leyenda-chip {clase_oculto}" id="chip-{i}" data-symbol="{symbol}" onclick="toggleMoneda({i})">'
            f'<span class="chip-punto" style="background:{color}"></span>{etiqueta}'
            f'<button class="chip-aislar" onclick="event.stopPropagation(); aislarMoneda({i})" '
            f'title="Ver solo esta moneda">🎯</button></span>'
        )

    labels_js = "[" + ",".join(f'"{ts}"' for ts in timestamps_grafico) + "]"
    datasets_str = ",\n".join(datasets_js)
    chips_str = "".join(chips_html)

    html = f"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<title>Binance Futures - Top Ganadoras (Dashboard)</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4"></script>
<style>
    body {{ font-family: Arial, sans-serif; margin: 20px; background: #f5f5f5; }}
    h2 {{ color: #222; }}
    h3 {{ color: #222; margin-top: 0; }}
    p {{ color: #555; }}
    #chartContainer {{ background: white; padding: 20px; border-radius: 8px; box-shadow: 0 1px 4px rgba(0,0,0,0.15); }}
    .alertas-caja {{ background: white; padding: 16px 20px; border-radius: 8px; box-shadow: 0 1px 4px rgba(0,0,0,0.15); margin-bottom: 16px; overflow-x: auto; }}
    .alertas-activas {{ border-left: 4px solid #ff6b35; }}
    .sin-alertas {{ color: #888; font-style: italic; margin: 4px 0; }}
    .nota-analisis {{ font-size: 13px; color: #666; background: #fff8e6; padding: 10px 14px;
                      border-radius: 6px; margin-bottom: 12px; line-height: 1.6; }}
    .veredicto-fav {{ color: #1a8917; font-weight: 700; cursor: help; }}
    .veredicto-neu {{ color: #b8860b; font-weight: 700; cursor: help; }}
    .veredicto-riesgo {{ color: #d32f2f; font-weight: 700; cursor: help; }}
    table {{ border-collapse: collapse; width: 100%; font-size: 14px; }}
    th, td {{ text-align: left; padding: 6px 10px; border-bottom: 1px solid #eee; white-space: nowrap; }}
    th {{ color: #777; font-weight: 600; }}
    .delta-pos {{ color: #1a8917; font-weight: 600; }}
    .delta-neg {{ color: #d32f2f; font-weight: 600; }}
    .tabs {{ margin-bottom: 0; }}
    .tab-btn {{ padding: 10px 20px; border: none; background: #e0e0e0; color: #555; cursor: pointer;
                border-radius: 8px 8px 0 0; margin-right: 4px; font-size: 14px; font-family: inherit; }}
    .tab-btn.active {{ background: white; color: #222; font-weight: bold; box-shadow: 0 -2px 4px rgba(0,0,0,0.08); }}
    .tab-content {{ display: none; }}
    .tab-content.active {{ display: block; }}
    .leyenda-chips {{ display: flex; flex-wrap: wrap; gap: 6px; margin-bottom: 12px; }}
    .leyenda-chip {{ display: inline-flex; align-items: center; gap: 5px; background: white;
                     border: 1px solid #ddd; border-radius: 16px; padding: 4px 6px 4px 10px;
                     font-size: 13px; cursor: pointer; user-select: none; color: #333; }}
    .leyenda-chip.chip-oculto {{ opacity: 0.4; text-decoration: line-through; }}
    .chip-punto {{ width: 10px; height: 10px; border-radius: 50%; display: inline-block; flex-shrink: 0; }}
    .chip-aislar {{ border: none; background: #f0f0f0; border-radius: 50%; width: 22px; height: 22px;
                    cursor: pointer; font-size: 12px; line-height: 1; padding: 0; }}
    .chip-aislar:hover {{ background: #ffdca8; }}
    .leyenda-acciones {{ margin-bottom: 8px; }}
    .btn-mostrar-todas {{ border: 1px solid #ccc; background: white; border-radius: 6px; padding: 6px 14px;
                          font-size: 13px; cursor: pointer; color: #333; }}
    .buscador-caja {{ margin-bottom: 10px; }}
    .buscador-caja input {{ width: 100%; max-width: 320px; padding: 8px 12px; border: 1px solid #ccc;
                            border-radius: 6px; font-size: 14px; box-sizing: border-box; }}
</style>
</head>
<body>
<h2>Binance Futures — % Cambio 24h por moneda</h2>
<p>Mostrando los últimos {VENTANA_DASHBOARD_DIAS} días. Pasa el mouse sobre una línea para resaltarla y ver a qué moneda corresponde. Clic en una etiqueta de abajo para mostrarla/ocultarla; clic en el 🎯 para ver solo esa moneda. Por defecto solo se muestran las del top actual o en racha. 🔥 = lleva {VENTANA_RACHA} corridas seguidas moviéndose fuerte y sostenido en la misma dirección. Última actualización: {timestamps_ordenados[-1]}</p>
<div class="tabs">
    <button class="tab-btn active" onclick="mostrarTab('grafico', this)">📈 Gráfico</button>
    <button class="tab-btn" onclick="mostrarTab('alertas', this)">🔔 Alertas{f' ({len(calificados)})' if calificados else ''}</button>
    <button class="tab-btn" onclick="mostrarTab('analisis', this)">📊 Análisis</button>
</div>
<div id="tab-grafico" class="tab-content active">
    <div class="buscador-caja">
        <input type="text" id="buscadorMoneda" placeholder="🔍 Buscar moneda..." oninput="filtrarChipsDebounced()">
    </div>
    <div class="leyenda-acciones">
        <button class="btn-mostrar-todas" onclick="mostrarTodas()">Mostrar todas</button>
        <button class="btn-mostrar-todas" onclick="ocultarTodas()">Ocultar todas</button>
    </div>
    <div class="leyenda-chips">{chips_str}</div>
    <div id="chartContainer">
    <canvas id="grafico" height="100"></canvas>
    </div>
</div>
<div id="tab-alertas" class="tab-content">
{html_alertas_activas}
{html_alertas_historial}
</div>
<div id="tab-analisis" class="tab-content">
{html_analisis}
</div>
<script>
function mostrarTab(nombre, boton) {{
    document.querySelectorAll('.tab-content').forEach(el => el.classList.remove('active'));
    document.querySelectorAll('.tab-btn').forEach(el => el.classList.remove('active'));
    document.getElementById('tab-' + nombre).classList.add('active');
    boton.classList.add('active');
}}
const ctx = document.getElementById('grafico').getContext('2d');
let ultimoHoverIdx = undefined;
const miChart = new Chart(ctx, {{
    type: 'line',
    data: {{
        labels: {labels_js},
        datasets: [
{datasets_str}
        ]
    }},
    options: {{
        responsive: true,
        animation: false,
        interaction: {{ mode: 'nearest', intersect: false }},
        onHover: (event, activeElements, chart) => {{
            const idx = activeElements.length ? activeElements[0].datasetIndex : null;
            if (idx === ultimoHoverIdx) return;  // evita redibujar si sigue sobre la misma línea
            ultimoHoverIdx = idx;
            chart.data.datasets.forEach((ds, i) => {{
                ds.borderWidth = (idx === null) ? 2 : (i === idx ? 4 : 1);
            }});
            chart.update('none');
        }},
        plugins: {{
            legend: {{ display: false }},
            tooltip: {{
                mode: 'nearest',
                intersect: false,
                callbacks: {{
                    title: (items) => items.length ? items[0].dataset.label : '',
                    label: (item) => `${{item.formattedValue}}%`
                }}
            }}
        }},
        scales: {{
            x: {{ title: {{ display: true, text: 'Fecha/Hora' }} }},
            y: {{ title: {{ display: true, text: '% Cambio 24h' }} }}
        }}
    }}
}});

function actualizarEstiloChips() {{
    miChart.data.datasets.forEach((ds, i) => {{
        const chip = document.getElementById('chip-' + i);
        if (!chip) return;
        chip.classList.toggle('chip-oculto', !!miChart.getDatasetMeta(i).hidden);
    }});
}}

function toggleMoneda(i) {{
    const meta = miChart.getDatasetMeta(i);
    meta.hidden = !meta.hidden;
    miChart.update();
    actualizarEstiloChips();
}}

function aislarMoneda(i) {{
    miChart.data.datasets.forEach((ds, idx) => {{
        miChart.getDatasetMeta(idx).hidden = (idx !== i);
    }});
    miChart.update();
    actualizarEstiloChips();
}}

function mostrarTodas() {{
    miChart.data.datasets.forEach((ds, idx) => {{
        miChart.getDatasetMeta(idx).hidden = false;
    }});
    miChart.update();
    actualizarEstiloChips();
}}

function ocultarTodas() {{
    miChart.data.datasets.forEach((ds, idx) => {{
        miChart.getDatasetMeta(idx).hidden = true;
    }});
    miChart.update();
    actualizarEstiloChips();
}}

function filtrarChips() {{
    const texto = document.getElementById('buscadorMoneda').value.toUpperCase();
    const visibles = [];
    document.querySelectorAll('.leyenda-chip').forEach(chip => {{
        const symbol = chip.getAttribute('data-symbol') || '';
        const coincide = symbol.toUpperCase().includes(texto);
        chip.style.display = coincide ? 'inline-flex' : 'none';
        if (coincide) visibles.push(chip);
    }});
    // si el buscador deja una sola moneda visible, la aísla sola en el gráfico
    if (texto.length > 0 && visibles.length === 1) {{
        const idx = parseInt(visibles[0].id.replace('chip-', ''));
        aislarMoneda(idx);
    }}
}}

let buscadorTimeoutId = null;
function filtrarChipsDebounced() {{
    // espera un instante corto después de dejar de escribir antes de filtrar,
    // para no recalcular/redibujar en cada tecla mientras se escribe rápido
    clearTimeout(buscadorTimeoutId);
    buscadorTimeoutId = setTimeout(filtrarChips, 200);
}}
</script>
</body>
</html>
"""

    html_path = os.path.join(os.path.dirname(EXCEL_FILE), REPO_GITHUB_NOMBRE_ARCHIVO)
    with open(html_path, "w", encoding="utf-8") as f:
        f.write(html)

    publicar_en_github(html_path)


def publicar_en_github(html_path_local):
    """Copia el dashboard HTML a la carpeta del repositorio de GitHub y lo sube
    (git add + commit + push), para que quede disponible en la URL pública de
    GitHub Pages. No interrumpe el script si falla (ej. sin internet en ese
    momento): solo lo avisa por consola y la corrida sigue siendo válida en
    el Excel/HTML local.

    Si el script está corriendo DENTRO de GitHub Actions (nube), este paso se
    omite: ahí el propio workflow ya se encarga de hacer commit y push de
    todos los archivos generados (Excel + HTML) en un solo paso, así que
    hacerlo aquí también sería redundante."""
    if os.environ.get("GITHUB_ACTIONS") == "true":
        return

    if not PUBLICAR_EN_GITHUB:
        return

    if not os.path.isdir(REPO_GITHUB_CARPETA):
        print(f"Aviso: no se encontró la carpeta del repo de GitHub en {REPO_GITHUB_CARPETA}. "
              f"Se omite la publicación online.")
        return

    try:
        destino = os.path.join(REPO_GITHUB_CARPETA, REPO_GITHUB_NOMBRE_ARCHIVO)
        shutil.copyfile(html_path_local, destino)

        subprocess.run(
            ["git", "add", REPO_GITHUB_NOMBRE_ARCHIVO],
            cwd=REPO_GITHUB_CARPETA, check=True, capture_output=True, text=True, timeout=30
        )

        commit = subprocess.run(
            ["git", "commit", "-m", f"Actualización automática {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"],
            cwd=REPO_GITHUB_CARPETA, capture_output=True, text=True, timeout=30
        )
        # Si no hubo cambios desde la última corrida, git commit "falla" con un
        # mensaje de "nothing to commit" — no es un error real, se ignora.
        if commit.returncode != 0 and "nothing to commit" not in commit.stdout.lower():
            print(f"Aviso: git commit devolvió un problema: {commit.stdout} {commit.stderr}")
            return

        subprocess.run(
            ["git", "push"],
            cwd=REPO_GITHUB_CARPETA, check=True, capture_output=True, text=True, timeout=60
        )
        print(f"Dashboard publicado en GitHub Pages ({REPO_GITHUB_NOMBRE_ARCHIVO}).")

    except subprocess.TimeoutExpired:
        print("Aviso: la publicación en GitHub tardó demasiado (¿sin internet?). Se omite esta vez.")
    except subprocess.CalledProcessError as e:
        print(f"Aviso: no se pudo publicar en GitHub. {e.stderr if e.stderr else e}")
    except Exception as e:
        print(f"Aviso: error inesperado al publicar en GitHub: {e}")


def main():
    top_ganadoras = obtener_top_ganadoras()
    if not top_ganadoras:
        print("No se encontraron datos (revisa tu conexión o los filtros de volumen).")
        return
    guardar_en_excel(top_ganadoras)


if __name__ == "__main__":
    main()

# ============================================================
# CÓMO PROGRAMARLO CADA HORA EN WINDOWS (Task Scheduler)
# ============================================================
# 1. Instala Python (python.org) si no lo tienes, y marca
#    "Add Python to PATH" durante la instalación.
# 2. Abre CMD y corre:
#       pip install requests openpyxl matplotlib
# 3. Guarda este archivo en una carpeta fija, ej:
#       C:\BinanceTracker\binance_tracker.py
# 4. Abre "Programador de tareas" (Task Scheduler) en Windows.
# 5. Crear tarea básica:
#       - Nombre: Binance Tracker
#       - Desencadenador: Diariamente, repetir cada 1 hora,
#         durante 24 horas (o el rango que quieras)
#       - Acción: Iniciar un programa
#           Programa: python
#           Argumentos: "C:\BinanceTracker\binance_tracker.py"
#           Iniciar en: C:\BinanceTracker
# 6. Guarda. El archivo binance_gainers.xlsx se irá actualizando
#    solo cada hora, con una hoja "Grafica" que se regenera
#    automáticamente en cada corrida.
# ============================================================
